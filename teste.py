import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ====================================
# CREDENCIAIS OMIE
# ====================================

app_key = "2543276123388"
app_secret = "cd84271c41f00486e438191c09b49522"

# ====================================
# URL API OMIE
# ====================================

url = "https://app.omie.com.br/api/v1/geral/clientes/"

# ====================================
# CABEÇALHOS DAS 65 COLUNAS
# ====================================

cabecalhos = [""] * 65

cabecalhos[1] = "CNPJ/CPF"
cabecalhos[2] = "Razão Social"
cabecalhos[17] = "Website"
cabecalhos[18] = "Banco"
cabecalhos[19] = "Agência"
cabecalhos[20] = "Conta Corrente"
cabecalhos[22] = "Nome Titular Conta"
cabecalhos[64] = "Chave PIX"

# ====================================
# LISTA FINAL
# ====================================

linhas = []

pagina = 1

while True:

    payload = {
        "call": "ListarClientes",
        "app_key": app_key,
        "app_secret": app_secret,
        "param": [
            {
                "pagina": pagina,
                "registros_por_pagina": 500,
                "apenas_importado_api": "N"
            }
        ]
    }

    response = requests.post(url, json=payload)

    dados = response.json()

    clientes = dados.get("clientes_cadastro", [])

    if not clientes:
        break

    for cliente in clientes:

        # cria linha vazia com 65 colunas
        linha = [""] * 65

        dados_bancarios = cliente.get("dadosBancarios", {})

        # ====================================
        # PREENCHER APENAS COLUNAS NECESSÁRIAS
        # ====================================

        linha[1] = cliente.get("cnpj_cpf") or "N/D"
        linha[2] = cliente.get("razao_social") or "N/D"
        linha[17] = cliente.get("website") or "N/D"
        linha[18] = dados_bancarios.get("codigo_banco") or "N/D"
        linha[19] = dados_bancarios.get("agencia") or "N/D"
        linha[20] = dados_bancarios.get("conta_corrente") or "N/D"
        linha[22] = dados_bancarios.get("nome_titular") or "N/D"
        linha[64] = dados_bancarios.get("cChavePix") or "N/D"

        linhas.append(linha)

    print(f"Página {pagina} carregada...")

    pagina += 1

# ====================================
# TRANSFORMAR EM DATAFRAME
# ====================================

df = pd.DataFrame(linhas)

# ====================================
# ABRIR PLANILHA MODELO
# ====================================

wb = load_workbook("modelo.xlsx")

# ====================================
# ABA BASE OMIE
# ====================================

ws = wb["BASE OMIE"]

# ====================================
# VISÍVEL PARA TESTE
# ====================================

ws.sheet_state = "visible"

# ====================================
# LIMPAR DADOS ANTIGOS
# ====================================

ws.delete_rows(1, ws.max_row)

# ====================================
# INSERIR CABEÇALHOS
# ====================================

ws.append(cabecalhos)

# ====================================
# COLAR DADOS
# ====================================

for row in dataframe_to_rows(df, index=False, header=False):
    ws.append(row)

# ====================================
# SALVAR
# ====================================

wb.save("relatorio_final.xlsx")

print("Relatório final criado com sucesso!")