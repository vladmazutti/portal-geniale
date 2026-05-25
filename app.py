from flask import Flask, send_file, request, redirect, url_for, render_template, jsonify
import requests
from requests.exceptions import Timeout, RequestException
from openpyxl import load_workbook
from datetime import datetime
from zoneinfo import ZoneInfo
from apscheduler.schedulers.background import BackgroundScheduler
from threading import Thread, Lock
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import os
import json
import traceback
import zipfile
import shutil
import tempfile
import xml.etree.ElementTree as ET


app = Flask(__name__)

FUSO_BRASIL = ZoneInfo("America/Sao_Paulo")

LOG_DIR = "logs"
CACHE_DIR = "cache"

OMIE_REGISTROS_POR_PAGINA = 1000
OMIE_TENTATIVA_MAXIMA = 5
OMIE_TIMEOUT = 90
OMIE_MAX_WORKERS = 4

COLUNAS_BASE_OMIE = 65

jobs_em_execucao = {}
jobs_lock = Lock()


RELATORIOS = {
    "geniale": {
        "nome": "GENIALE",
        "titulo": "Planilha de Pagamento",
        "descricao": "Pagamentos operacionais Geniale",
        "modelo": "modelo.xlsx",
        "arquivo_pronto": "planilha_geniale.xlsx",
        "arquivo_atualizacao": "ultima_atualizacao_geniale.txt",
        "arquivo_status": "status_geniale.txt",
        "arquivo_log": "logs/geniale.log",
        "arquivo_cache": "cache/cache_omie_geniale.json",
        "env_key": "OMIE_APP_KEY",
        "env_secret": "OMIE_APP_SECRET",
        "download_name": "Pagamentos_Geniale",
        "restrito": False,
        "logo": "logo.png",
        "classe": "geniale",
    },
    "paniz": {
        "nome": "PANIZ",
        "titulo": "Planilha de Pagamento",
        "descricao": "Pagamentos operacionais Paniz",
        "modelo": "modelo.xlsx",
        "arquivo_pronto": "planilha_paniz.xlsx",
        "arquivo_atualizacao": "ultima_atualizacao_paniz.txt",
        "arquivo_status": "status_paniz.txt",
        "arquivo_log": "logs/paniz.log",
        "arquivo_cache": "cache/cache_omie_paniz.json",
        "env_key": "OMIE_PANIZ_APP_KEY",
        "env_secret": "OMIE_PANIZ_APP_SECRET",
        "download_name": "Pagamentos_Paniz",
        "restrito": False,
        "logo": "logo_paniz.png",
        "classe": "paniz",
    },
    "financeiro": {
        "nome": "FINANCEIRO",
        "titulo": "Consolidadora Financeiro",
        "descricao": "Planilha consolidada de acesso restrito",
        "modelo": "modelo_financeiro.xlsx",
        "arquivo_pronto": "planilha_financeiro.xlsx",
        "arquivo_atualizacao": "ultima_atualizacao_financeiro.txt",
        "arquivo_status": "status_financeiro.txt",
        "arquivo_log": "logs/financeiro.log",
        "arquivo_cache": "cache/cache_omie_financeiro.json",
        "env_key": "OMIE_APP_KEY",
        "env_secret": "OMIE_APP_SECRET",
        "download_name": "Consolidadora_Financeiro",
        "restrito": True,
        "logo": None,
        "classe": "financeiro",
    },
}


def garantir_pasta_logs():
    if not os.path.exists(LOG_DIR):
        os.makedirs(LOG_DIR)


def garantir_pasta_cache():
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)


def agora_brasil():
    return datetime.now(FUSO_BRASIL)


def agora_formatado():
    return agora_brasil().strftime("%d/%m/%Y %H:%M:%S")


def data_arquivo():
    return agora_brasil().strftime("%d%m%y")


def formatar_duracao(segundos):
    segundos = int(segundos)
    minutos = segundos // 60
    segundos_restantes = segundos % 60

    if minutos <= 0:
        return f"{segundos_restantes}s"

    return f"{minutos}m {segundos_restantes}s"


def escrever_arquivo(caminho, conteudo):
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(conteudo)


def adicionar_ao_arquivo(caminho, conteudo):
    garantir_pasta_logs()

    with open(caminho, "a", encoding="utf-8") as f:
        f.write(conteudo)


def ler_arquivo(caminho, padrao=""):
    if os.path.exists(caminho):
        with open(caminho, "r", encoding="utf-8") as f:
            return f.read().strip()

    return padrao


def definir_status(tipo, status):
    config = RELATORIOS[tipo]
    escrever_arquivo(config["arquivo_status"], status)


def obter_ultima_atualizacao(config):
    return ler_arquivo(
        config["arquivo_atualizacao"],
        "Ainda não atualizada"
    )


def obter_ultimo_log(config):
    caminho = config.get("arquivo_log")

    if not caminho or not os.path.exists(caminho):
        return "Sem histórico de atualização"

    conteudo = ler_arquivo(caminho, "")

    if not conteudo:
        return "Sem histórico de atualização"

    blocos = conteudo.split("\n\n")

    for bloco in reversed(blocos):
        bloco = bloco.strip()

        if bloco:
            return bloco

    return "Sem histórico de atualização"


def registrar_log_atualizacao(
    tipo,
    status,
    inicio,
    fim,
    registros=0,
    paginas=0,
    modo="Paralelo",
    cache="Snapshot atualizado",
    erro=""
):
    config = RELATORIOS[tipo]

    duracao_segundos = (fim - inicio).total_seconds()
    duracao_formatada = formatar_duracao(duracao_segundos)

    linhas_log = [
        "============================================================",
        f"DATA/HORA: {fim.strftime('%d/%m/%Y %H:%M:%S')}",
        f"RELATÓRIO: {config['nome']}",
        f"STATUS: {status}",
        f"INÍCIO: {inicio.strftime('%d/%m/%Y %H:%M:%S')}",
        f"FIM: {fim.strftime('%d/%m/%Y %H:%M:%S')}",
        f"DURAÇÃO: {duracao_formatada}",
        f"REGISTROS OMIE: {registros}",
        f"PÁGINAS OMIE: {paginas}",
        f"MODO OMIE: {modo}",
        f"WORKERS OMIE: {OMIE_MAX_WORKERS}",
        f"CACHE: {cache}",
    ]

    if erro:
        linhas_log.append(f"ERRO: {erro}")

    linhas_log.append("============================================================")
    linhas_log.append("")
    linhas_log.append("")

    adicionar_ao_arquivo(
        config["arquivo_log"],
        "\n".join(linhas_log)
    )


def obter_status(tipo, config):
    status_interno = ler_arquivo(config["arquivo_status"], "")

    if status_interno == "atualizando":
        return {
            "texto": "Atualizando...",
            "cor": "atualizando",
            "existe": os.path.exists(config["arquivo_pronto"]),
            "ultima": obter_ultima_atualizacao(config),
            "ultimo_log": obter_ultimo_log(config),
        }

    if status_interno.startswith("erro"):
        return {
            "texto": "Erro na atualização",
            "cor": "erro",
            "existe": os.path.exists(config["arquivo_pronto"]),
            "ultima": obter_ultima_atualizacao(config),
            "ultimo_log": obter_ultimo_log(config),
        }

    existe = os.path.exists(config["arquivo_pronto"])

    return {
        "texto": "Disponível para download" if existe else "Ainda não gerada",
        "cor": "ok" if existe else "pendente",
        "existe": existe,
        "ultima": obter_ultima_atualizacao(config),
        "ultimo_log": obter_ultimo_log(config),
    }


def montar_relatorios_para_tela():
    lista = []

    for tipo, config in RELATORIOS.items():
        item = dict(config)
        item["tipo"] = tipo
        item["status"] = obter_status(tipo, config)
        lista.append(item)

    return lista


def buscar_pagina_omie(
    app_key,
    app_secret,
    pagina,
    tentativa_maxima=OMIE_TENTATIVA_MAXIMA
):
    url = "https://app.omie.com.br/api/v1/geral/clientes/"

    payload = {
        "call": "ListarClientes",
        "app_key": app_key,
        "app_secret": app_secret,
        "param": [
            {
                "pagina": pagina,
                "registros_por_pagina": OMIE_REGISTROS_POR_PAGINA,
                "apenas_importado_api": "N",
            }
        ],
    }

    for tentativa in range(1, tentativa_maxima + 1):
        try:
            response = requests.post(
                url,
                json=payload,
                timeout=OMIE_TIMEOUT
            )

            if response.status_code == 429:
                espera = tentativa * 15

                print(
                    f"Limite OMIE atingido na página {pagina}. "
                    f"Tentativa {tentativa}/{tentativa_maxima}. "
                    f"Aguardando {espera}s..."
                )

                time.sleep(espera)
                continue

            response.raise_for_status()

            return response.json()

        except Timeout:
            espera = tentativa * 8

            print(
                f"Timeout na página {pagina}. "
                f"Tentativa {tentativa}/{tentativa_maxima}. "
                f"Aguardando {espera}s..."
            )

            time.sleep(espera)

        except RequestException as erro:
            espera = tentativa * 8

            print(
                f"Erro de conexão na página {pagina}: {erro}. "
                f"Tentativa {tentativa}/{tentativa_maxima}. "
                f"Aguardando {espera}s..."
            )

            time.sleep(espera)

    raise Exception(
        f"Falha ao consultar OMIE na página {pagina} "
        f"após {tentativa_maxima} tentativas."
    )


def montar_linha_cliente(cliente):
    linha = [""] * COLUNAS_BASE_OMIE
    dados_bancarios = cliente.get("dadosBancarios", {})

    linha[1] = cliente.get("cnpj_cpf") or "N/D"
    linha[2] = cliente.get("razao_social") or "N/D"
    linha[17] = cliente.get("website") or "N/D"
    linha[18] = dados_bancarios.get("codigo_banco") or "N/D"
    linha[19] = dados_bancarios.get("agencia") or "N/D"
    linha[20] = dados_bancarios.get("conta_corrente") or "N/D"
    linha[22] = dados_bancarios.get("nome_titular") or "N/D"
    linha[64] = dados_bancarios.get("cChavePix") or "N/D"

    return linha


def processar_clientes_pagina(dados):
    linhas = []
    clientes = dados.get("clientes_cadastro", [])

    for cliente in clientes:
        linhas.append(montar_linha_cliente(cliente))

    return linhas


def consultar_base_omie_completa(app_key, app_secret):
    print("Consultando primeira página OMIE para identificar total de páginas...")

    dados_primeira_pagina = buscar_pagina_omie(
        app_key,
        app_secret,
        1
    )

    total_paginas = dados_primeira_pagina.get("total_de_paginas", 1)

    print(f"Total de páginas OMIE identificado: {total_paginas}")

    paginas_resultado = {
        1: processar_clientes_pagina(dados_primeira_pagina)
    }

    if total_paginas <= 1:
        linhas = paginas_resultado[1]

        return {
            "linhas": linhas,
            "total_paginas": total_paginas,
            "modo": "Sequencial",
        }

    paginas_restantes = list(range(2, total_paginas + 1))

    print(
        f"Iniciando busca paralela OMIE com "
        f"{OMIE_MAX_WORKERS} workers..."
    )

    with ThreadPoolExecutor(max_workers=OMIE_MAX_WORKERS) as executor:
        tarefas = {}

        for pagina in paginas_restantes:
            tarefa = executor.submit(
                buscar_pagina_omie,
                app_key,
                app_secret,
                pagina
            )

            tarefas[tarefa] = pagina

        for tarefa in as_completed(tarefas):
            pagina = tarefas[tarefa]

            try:
                dados = tarefa.result()
                paginas_resultado[pagina] = processar_clientes_pagina(dados)

                print(f"Página {pagina}/{total_paginas} carregada...")

            except Exception as erro:
                raise Exception(
                    f"Erro ao processar página OMIE {pagina}: {erro}"
                )

    linhas = []

    for pagina in range(1, total_paginas + 1):
        linhas.extend(paginas_resultado.get(pagina, []))

    print(
        f"Busca OMIE completa concluída. "
        f"Páginas: {total_paginas}. "
        f"Registros: {len(linhas)}."
    )

    return {
        "linhas": linhas,
        "total_paginas": total_paginas,
        "modo": "Paralelo",
    }


def salvar_snapshot_cache(config, linhas, total_paginas, modo):
    garantir_pasta_cache()

    caminho = config["arquivo_cache"]
    arquivo_temporario = f"{caminho}.tmp"

    snapshot = {
        "gerado_em": agora_brasil().isoformat(),
        "tipo": "snapshot_completo",
        "total_registros": len(linhas),
        "total_paginas": total_paginas,
        "modo": modo,
        "observacao": (
            "Cache apenas informativo/fallback. "
            "A planilha é sempre gerada com consulta completa à OMIE."
        ),
        "linhas": linhas,
    }

    with open(arquivo_temporario, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False)

    os.replace(arquivo_temporario, caminho)


def buscar_clientes_omie(app_key, app_secret, config):
    resultado = consultar_base_omie_completa(
        app_key,
        app_secret
    )

    linhas = resultado["linhas"]
    total_paginas = resultado["total_paginas"]
    modo = resultado["modo"]

    salvar_snapshot_cache(
        config,
        linhas,
        total_paginas,
        modo
    )

    return {
        "linhas": linhas,
        "total_paginas": total_paginas,
        "modo": modo,
        "cache": "Snapshot completo atualizado",
    }


def limpar_base_omie_sem_deletar_linhas(ws, total_linhas_novas):
    ultima_linha_existente = max(ws.max_row, total_linhas_novas, 1)

    for numero_linha in range(1, ultima_linha_existente + 1):
        for numero_coluna in range(1, COLUNAS_BASE_OMIE + 1):
            ws.cell(
                row=numero_linha,
                column=numero_coluna
            ).value = None


def escrever_linha_base_omie(ws, numero_linha, valores):
    for indice, valor in enumerate(valores, start=1):
        ws.cell(
            row=numero_linha,
            column=indice
        ).value = valor


def montar_cabecalhos_base_omie():
    cabecalhos = [""] * COLUNAS_BASE_OMIE

    cabecalhos[1] = "CNPJ/CPF"
    cabecalhos[2] = "Razão Social"
    cabecalhos[17] = "Website"
    cabecalhos[18] = "Banco"
    cabecalhos[19] = "Agência"
    cabecalhos[20] = "Conta Corrente"
    cabecalhos[22] = "Nome Titular Conta"
    cabecalhos[64] = "Chave PIX"

    return cabecalhos


def atualizar_base_omie(ws, linhas):
    total_linhas_novas = len(linhas) + 1

    limpar_base_omie_sem_deletar_linhas(
        ws,
        total_linhas_novas
    )

    escrever_linha_base_omie(
        ws,
        1,
        montar_cabecalhos_base_omie()
    )

    numero_linha = 2

    for linha in linhas:
        escrever_linha_base_omie(
            ws,
            numero_linha,
            linha
        )

        numero_linha += 1


def remover_calcchain_do_xlsx(pasta_extraida):
    caminho_calcchain = os.path.join(
        pasta_extraida,
        "xl",
        "calcChain.xml"
    )

    if os.path.exists(caminho_calcchain):
        os.remove(caminho_calcchain)

    caminho_rels = os.path.join(
        pasta_extraida,
        "xl",
        "_rels",
        "workbook.xml.rels"
    )

    if os.path.exists(caminho_rels):
        tree = ET.parse(caminho_rels)
        root = tree.getroot()

        for rel in list(root):
            tipo = rel.attrib.get("Type", "")
            target = rel.attrib.get("Target", "")

            if "calcChain" in tipo or "calcChain" in target:
                root.remove(rel)

        tree.write(
            caminho_rels,
            encoding="utf-8",
            xml_declaration=True
        )

    caminho_content_types = os.path.join(
        pasta_extraida,
        "[Content_Types].xml"
    )

    if os.path.exists(caminho_content_types):
        tree = ET.parse(caminho_content_types)
        root = tree.getroot()

        for item in list(root):
            part_name = item.attrib.get("PartName", "")

            if part_name == "/xl/calcChain.xml":
                root.remove(item)

        tree.write(
            caminho_content_types,
            encoding="utf-8",
            xml_declaration=True
        )


def normalizar_formulas_matriciais_do_xlsx(pasta_extraida):
    pasta_worksheets = os.path.join(
        pasta_extraida,
        "xl",
        "worksheets"
    )

    if not os.path.exists(pasta_worksheets):
        return

    for nome_arquivo in os.listdir(pasta_worksheets):
        if not nome_arquivo.endswith(".xml"):
            continue

        caminho_sheet = os.path.join(
            pasta_worksheets,
            nome_arquivo
        )

        tree = ET.parse(caminho_sheet)
        root = tree.getroot()

        alterado = False

        for formula in root.iter():
            tag_limpa = formula.tag.split("}")[-1]

            if tag_limpa != "f":
                continue

            if formula.attrib.get("t") == "array":
                formula.attrib.pop("t", None)
                formula.attrib.pop("ref", None)
                formula.attrib.pop("aca", None)
                formula.attrib.pop("ca", None)
                formula.attrib.pop("si", None)
                alterado = True

        if alterado:
            tree.write(
                caminho_sheet,
                encoding="utf-8",
                xml_declaration=True
            )


def forcar_recalculo_excel(pasta_extraida):
    caminho_workbook = os.path.join(
        pasta_extraida,
        "xl",
        "workbook.xml"
    )

    if not os.path.exists(caminho_workbook):
        return

    tree = ET.parse(caminho_workbook)
    root = tree.getroot()

    namespace = ""

    if root.tag.startswith("{"):
        namespace = root.tag.split("}")[0].replace("{", "")

    if namespace:
        calc_pr = root.find(f"{{{namespace}}}calcPr")
    else:
        calc_pr = root.find("calcPr")

    if calc_pr is None:
        if namespace:
            calc_pr = ET.SubElement(root, f"{{{namespace}}}calcPr")
        else:
            calc_pr = ET.SubElement(root, "calcPr")

    calc_pr.set("calcMode", "auto")
    calc_pr.set("fullCalcOnLoad", "1")
    calc_pr.set("forceFullCalc", "1")

    tree.write(
        caminho_workbook,
        encoding="utf-8",
        xml_declaration=True
    )


def recomprimir_xlsx(pasta_extraida, arquivo_saida):
    with zipfile.ZipFile(
        arquivo_saida,
        "w",
        compression=zipfile.ZIP_DEFLATED
    ) as zip_saida:
        for raiz, _, arquivos in os.walk(pasta_extraida):
            for nome_arquivo in arquivos:
                caminho_absoluto = os.path.join(
                    raiz,
                    nome_arquivo
                )

                caminho_relativo = os.path.relpath(
                    caminho_absoluto,
                    pasta_extraida
                )

                caminho_relativo = caminho_relativo.replace(
                    os.sep,
                    "/"
                )

                zip_saida.write(
                    caminho_absoluto,
                    caminho_relativo
                )


def sanitizar_xlsx_pos_salvamento(arquivo_xlsx):
    pasta_temporaria = tempfile.mkdtemp()

    try:
        with zipfile.ZipFile(arquivo_xlsx, "r") as zip_entrada:
            zip_entrada.extractall(pasta_temporaria)

        remover_calcchain_do_xlsx(pasta_temporaria)
        normalizar_formulas_matriciais_do_xlsx(pasta_temporaria)
        forcar_recalculo_excel(pasta_temporaria)

        arquivo_sanitizado = f"{arquivo_xlsx}.sanitized.xlsx"

        if os.path.exists(arquivo_sanitizado):
            os.remove(arquivo_sanitizado)

        recomprimir_xlsx(
            pasta_temporaria,
            arquivo_sanitizado
        )

        os.replace(
            arquivo_sanitizado,
            arquivo_xlsx
        )

    finally:
        shutil.rmtree(
            pasta_temporaria,
            ignore_errors=True
        )


def salvar_workbook_com_seguranca(wb, arquivo_final):
    arquivo_temporario = f"{arquivo_final}.tmp.xlsx"

    if os.path.exists(arquivo_temporario):
        os.remove(arquivo_temporario)

    wb.save(arquivo_temporario)

    if not os.path.exists(arquivo_temporario):
        raise Exception("Arquivo temporário não foi gerado corretamente.")

    sanitizar_xlsx_pos_salvamento(arquivo_temporario)

    os.replace(arquivo_temporario, arquivo_final)


def gerar_planilha(tipo):
    if tipo not in RELATORIOS:
        raise Exception("Relatório inválido.")

    config = RELATORIOS[tipo]

    print(f"Iniciando atualização da planilha {config['nome']}...")

    app_key = os.getenv(config["env_key"])
    app_secret = os.getenv(config["env_secret"])

    if not app_key or not app_secret:
        raise Exception(
            f"Credenciais OMIE não configuradas para {config['nome']}. "
            f"Verifique {config['env_key']} e {config['env_secret']} no Railway."
        )

    if not os.path.exists(config["modelo"]):
        raise Exception(f"Modelo não encontrado: {config['modelo']}")

    resultado_omie = buscar_clientes_omie(
        app_key,
        app_secret,
        config
    )

    linhas = resultado_omie["linhas"]
    total_paginas = resultado_omie["total_paginas"]
    modo = resultado_omie["modo"]
    cache_status = resultado_omie["cache"]

    wb = load_workbook(config["modelo"])

    if "BASE OMIE" not in wb.sheetnames:
        raise Exception(f"Aba 'BASE OMIE' não encontrada em {config['modelo']}")

    ws = wb["BASE OMIE"]
    ws.sheet_state = "veryHidden"

    atualizar_base_omie(
        ws,
        linhas
    )

    salvar_workbook_com_seguranca(
        wb,
        config["arquivo_pronto"]
    )

    escrever_arquivo(
        config["arquivo_atualizacao"],
        agora_formatado()
    )

    print(f"Planilha {config['nome']} atualizada com sucesso!")

    return {
        "registros": len(linhas),
        "paginas": total_paginas,
        "modo": modo,
        "cache": cache_status,
    }


def executar_atualizacao_background(tipo):
    inicio = agora_brasil()
    registros = 0
    paginas = 0
    modo = "Paralelo"
    cache_status = "Snapshot completo"

    try:
        definir_status(tipo, "atualizando")

        resultado = gerar_planilha(tipo)

        registros = resultado.get("registros", 0)
        paginas = resultado.get("paginas", 0)
        modo = resultado.get("modo", "Paralelo")
        cache_status = resultado.get("cache", "Snapshot completo")

        definir_status(tipo, "ok")

        fim = agora_brasil()

        registrar_log_atualizacao(
            tipo=tipo,
            status="OK",
            inicio=inicio,
            fim=fim,
            registros=registros,
            paginas=paginas,
            modo=modo,
            cache=cache_status,
            erro=""
        )

    except Exception as erro:
        fim = agora_brasil()
        erro_texto = str(erro)

        print(f"Erro ao atualizar {tipo}: {erro_texto}")
        print(traceback.format_exc())

        definir_status(tipo, f"erro: {erro_texto}")

        registrar_log_atualizacao(
            tipo=tipo,
            status="ERRO",
            inicio=inicio,
            fim=fim,
            registros=registros,
            paginas=paginas,
            modo=modo,
            cache=cache_status,
            erro=erro_texto
        )

    finally:
        with jobs_lock:
            jobs_em_execucao[tipo] = False


def iniciar_atualizacao_background(tipo):
    if tipo not in RELATORIOS:
        raise Exception("Relatório inválido.")

    with jobs_lock:
        if jobs_em_execucao.get(tipo):
            return False

        jobs_em_execucao[tipo] = True

    thread = Thread(
        target=executar_atualizacao_background,
        args=(tipo,),
        daemon=True
    )
    thread.start()

    return True


def atualizar_agendado(tipo):
    try:
        iniciar_atualizacao_background(tipo)
    except Exception as erro:
        print(f"Erro ao iniciar atualização agendada de {tipo}: {erro}")


scheduler = BackgroundScheduler(timezone="America/Sao_Paulo")

scheduler.add_job(
    lambda: atualizar_agendado("geniale"),
    trigger="cron",
    hour=8,
    minute=0,
)

scheduler.add_job(
    lambda: atualizar_agendado("paniz"),
    trigger="cron",
    hour=8,
    minute=0,
)

scheduler.add_job(
    lambda: atualizar_agendado("financeiro"),
    trigger="cron",
    hour="8,12",
    minute=0,
)

scheduler.start()


@app.route("/")
def home():
    return render_template(
        "index.html",
        relatorios=montar_relatorios_para_tela(),
        agora=agora_formatado(),
    )


@app.route("/api/status")
def api_status():
    return jsonify(montar_relatorios_para_tela())


@app.route("/api/atualizar/<tipo>")
def api_atualizar(tipo):
    if tipo not in RELATORIOS:
        return jsonify({
            "sucesso": False,
            "mensagem": "Relatório inválido."
        }), 404

    try:
        iniciado = iniciar_atualizacao_background(tipo)

        if iniciado:
            return jsonify({
                "sucesso": True,
                "mensagem": (
                    f"A atualização da planilha {RELATORIOS[tipo]['nome']} "
                    f"foi iniciada em segundo plano."
                )
            })

        return jsonify({
            "sucesso": False,
            "mensagem": (
                f"A planilha {RELATORIOS[tipo]['nome']} "
                f"já está sendo atualizada."
            )
        })

    except Exception as erro:
        return jsonify({
            "sucesso": False,
            "mensagem": str(erro)
        }), 500


@app.route("/baixar/<tipo>")
def baixar(tipo):
    if tipo not in RELATORIOS:
        return render_template(
            "erro.html",
            titulo="Relatório inválido",
            mensagem="O relatório solicitado não existe.",
        ), 404

    config = RELATORIOS[tipo]

    if config["restrito"]:
        return redirect(url_for("acessar", tipo=tipo))

    if not os.path.exists(config["arquivo_pronto"]):
        return render_template(
            "erro.html",
            titulo="A planilha ainda não foi gerada.",
            mensagem="Clique em Atualizar para gerar a primeira versão.",
        )

    return send_file(
        config["arquivo_pronto"],
        as_attachment=True,
        download_name=f"{config['download_name']}_{data_arquivo()}.xlsx",
    )


@app.route("/acessar/<tipo>", methods=["GET", "POST"])
def acessar(tipo):
    if tipo not in RELATORIOS:
        return render_template(
            "erro.html",
            titulo="Relatório inválido",
            mensagem="O relatório solicitado não existe.",
        ), 404

    config = RELATORIOS[tipo]

    if not config["restrito"]:
        return redirect(url_for("baixar", tipo=tipo))

    erro = ""

    if request.method == "POST":
        senha_digitada = request.form.get("senha", "")
        senha_correta = os.getenv("FINANCEIRO_PASSWORD")

        if not senha_correta:
            erro = "Senha do Financeiro não configurada no Railway."

        elif senha_digitada == senha_correta:
            if not os.path.exists(config["arquivo_pronto"]):
                return render_template(
                    "erro.html",
                    titulo="A planilha ainda não foi gerada.",
                    mensagem="Clique em Atualizar para gerar a primeira versão.",
                )

            return send_file(
                config["arquivo_pronto"],
                as_attachment=True,
                download_name=f"{config['download_name']}_{data_arquivo()}.xlsx",
            )

        else:
            erro = "Senha incorreta."

    return render_template(
        "acesso.html",
        tipo=tipo,
        relatorio=config,
        erro=erro,
    )


@app.route("/atualizar/<tipo>")
def atualizar(tipo):
    if tipo not in RELATORIOS:
        return render_template(
            "erro.html",
            titulo="Relatório inválido",
            mensagem="O relatório solicitado não existe.",
        ), 404

    try:
        iniciado = iniciar_atualizacao_background(tipo)

    except Exception as erro:
        return render_template(
            "erro.html",
            titulo="Erro ao iniciar atualização.",
            mensagem=str(erro),
        )

    if iniciado:
        mensagem = (
            f"A atualização da planilha {RELATORIOS[tipo]['nome']} "
            f"foi iniciada em segundo plano."
        )
    else:
        mensagem = (
            f"A planilha {RELATORIOS[tipo]['nome']} "
            f"já está sendo atualizada."
        )

    return render_template(
        "sucesso.html",
        relatorio=RELATORIOS[tipo],
        mensagem=mensagem,
    )


garantir_pasta_logs()
garantir_pasta_cache()


if __name__ == "__main__":
    app.run(debug=True)